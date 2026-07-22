"""Formatted Excel report for the salary-after-rent analysis.

Builds a management-friendly workbook (openpyxl) from the DataFrame produced by
analysis.analyze_period():
  - "Summary"      country ranking with averages, conditional formatting and a bar chart
  - "By Year"      country x year matrix of disposable income (intl $)
  - "Methodology"  formula, data sources and main caveats

Used via:  python analysis.py --case 6 --start-year 2018 --save-plots --excel
"""

import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from etl import FULL_COUNTRY_NAMES

HEADER_FILL = PatternFill('solid', fgColor='2C3E50')
HEADER_FONT = Font(bold=True, color='FFFFFF')
TITLE_FONT = Font(bold=True, size=14)
NEGATIVE_FILL = PatternFill('solid', fgColor='FADBD8')
NEGATIVE_FONT = Font(color='943126')
INT_FORMAT = '#,##0'


def _style_header_row(ws, row, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _autofit(ws, widths):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _summary_table(merged_df):
    """Per-country averages over the analysis period, sorted best-first."""
    grouped = merged_df.groupby('Country Code').agg(
        salary_EUR=('OBS_VALUE', 'mean'),
        rent_EUR=('rent_EUR', 'mean'),
        disposable_EUR=('salary_minus_housing_EUR', 'mean'),
        disposable_intl=('salary_minus_housing', 'mean'),
        year_min=('TIME_PERIOD', 'min'),
        year_max=('TIME_PERIOD', 'max'),
        n_years=('TIME_PERIOD', 'count'),
    ).sort_values('disposable_intl', ascending=False).reset_index()
    grouped['Country'] = grouped['Country Code'].map(FULL_COUNTRY_NAMES).fillna(grouped['Country Code'])
    return grouped


def _build_summary_sheet(ws, summary, earnings_case, start_year):
    ws.title = 'Summary'
    ws['A1'] = 'European Salary Analysis - Disposable Income After Rent (PPP-adjusted)'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'Earnings case: {earnings_case}'
    ws['A3'] = (f'Period: from {start_year} | Averages per country | '
                f'Generated: {datetime.date.today().isoformat()}')
    for row in (2, 3):
        ws.cell(row=row, column=1).font = Font(italic=True, color='555555')

    header_row = 5
    headers = ['Rank', 'Country', 'Avg. net salary (EUR/year)', 'Avg. rent (EUR/month)',
               'Avg. disposable (EUR/year)', 'Avg. disposable (intl $/year)', 'Years covered']
    for col, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=header)
    _style_header_row(ws, header_row, len(headers))

    for i, row in summary.iterrows():
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=row['Country'])
        for col, key in ((3, 'salary_EUR'), (4, 'rent_EUR'), (5, 'disposable_EUR'), (6, 'disposable_intl')):
            cell = ws.cell(row=r, column=col, value=round(float(row[key])))
            cell.number_format = INT_FORMAT
        ws.cell(row=r, column=7,
                value=f"{int(row['year_min'])}-{int(row['year_max'])} ({int(row['n_years'])})")

    first_data, last_data = header_row + 1, header_row + len(summary)
    # Flag countries where average rent exceeds the net salary
    ws.conditional_formatting.add(
        f'E{first_data}:F{last_data}',
        CellIsRule(operator='lessThan', formula=['0'], fill=NEGATIVE_FILL, font=NEGATIVE_FONT))

    chart = BarChart()
    chart.type = 'bar'
    chart.title = 'Average disposable income after rent (intl $/year)'
    chart.legend = None
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    data = Reference(ws, min_col=6, min_row=first_data, max_row=last_data)
    cats = Reference(ws, min_col=2, min_row=first_data, max_row=last_data)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 0.55 * len(summary) + 3
    chart.width = 18
    ws.add_chart(chart, f'I{header_row}')

    ws.freeze_panes = ws.cell(row=first_data, column=1)
    _autofit(ws, [6, 18, 22, 18, 22, 24, 16])


def _build_by_year_sheet(ws, merged_df):
    ws.title = 'By Year'
    pivot = merged_df.pivot_table(index='Country Code', columns='TIME_PERIOD',
                                  values='salary_minus_housing', aggfunc='first')
    pivot.index = [FULL_COUNTRY_NAMES.get(c, c) for c in pivot.index]
    pivot = pivot.sort_index()

    ws.cell(row=1, column=1, value='Disposable income after rent, intl $/year')
    ws.cell(row=1, column=1).font = Font(bold=True)

    header_row = 3
    ws.cell(row=header_row, column=1, value='Country')
    years = [int(y) for y in pivot.columns]
    for col, year in enumerate(years, start=2):
        ws.cell(row=header_row, column=col, value=year)
    _style_header_row(ws, header_row, len(years) + 1)

    for i, (country, values) in enumerate(pivot.iterrows()):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=country)
        for col, year in enumerate(pivot.columns, start=2):
            value = values[year]
            if value == value:  # not NaN
                cell = ws.cell(row=r, column=col, value=round(float(value)))
                cell.number_format = INT_FORMAT

    first_data, last_data = header_row + 1, header_row + len(pivot)
    last_col = get_column_letter(len(years) + 1)
    ws.conditional_formatting.add(
        f'B{first_data}:{last_col}{last_data}',
        CellIsRule(operator='lessThan', formula=['0'], fill=NEGATIVE_FILL, font=NEGATIVE_FONT))

    ws.freeze_panes = 'B4'
    _autofit(ws, [18] + [10] * len(years))


METHODOLOGY_LINES = [
    ('European Salary Analysis - Methodology', True),
    ('', False),
    ('Formula:', True),
    ('Disposable income (intl $) = (annual net salary EUR - 12 x monthly rent EUR) / PPP factor (EUR per intl $)', False),
    ('', False),
    ('Data sources:', True),
    ('- Eurostat EARN_NT_NET: annual net earnings by country and household type (EUR), via the Eurostat SDMX API', False),
    ('- Eurostat PRC_COLC_RENTS: monthly rent, 1-bedroom flat in the capital city (EUR)', False),
    ('- World Bank PA.NUS.PPP: PPP conversion factor (LCU per international $)', False),
    ('- ECB eurofxref-hist: daily EUR reference exchange rates, averaged per year', False),
    ('', False),
    ('Harmonisation:', True),
    ('- Non-Eurozone countries: PPP (LCU/intl $) divided by the annual average LCU-per-EUR rate', False),
    ('- Countries that joined the Eurozone (HRV, LTU, LVA, MLT): World Bank PPP is already re-expressed in EUR '
     'and is NOT divided by historical exchange rates', False),
    ('', False),
    ('Main caveats:', True),
    ('- Rents are capital-city level, salaries are country-wide averages', False),
    ('- World Bank GDP PPP used instead of Eurostat CPL (5-20% possible difference)', False),
    ('- Confidence intervals in the charts reflect year-to-year variability, not sampling error', False),
]


def _build_methodology_sheet(ws):
    ws.title = 'Methodology'
    for row, (text, bold) in enumerate(METHODOLOGY_LINES, start=1):
        cell = ws.cell(row=row, column=1, value=text)
        if bold:
            cell.font = Font(bold=True)
    ws.column_dimensions['A'].width = 110


def build_excel_report(merged_df, earnings_case, start_year, out_path='output/salary_report.xlsx'):
    """Write the three-sheet formatted report; returns the output path."""
    wb = Workbook()
    _build_summary_sheet(wb.active, _summary_table(merged_df), earnings_case, start_year)
    _build_by_year_sheet(wb.create_sheet(), merged_df)
    _build_methodology_sheet(wb.create_sheet())
    wb.save(out_path)
    print(f"[INFO] Saved {out_path}")
    return out_path
